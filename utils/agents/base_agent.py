"""
utils/agents/base_agent.py

BaseAgent — parent class for all framework agents.
Handles:
  - AI provider selection (Anthropic preferred, Gemini fallback)
  - Claude API calls with system prompt support
  - File save utility
  - Excel / CSV / text input reading

All agents inherit from this. They never call the AI API directly.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv

load_dotenv()


class BaseAgent:
    """
    Parent class for all framework agents.

    Provider selection (automatic):
      ANTHROPIC_API_KEY set → uses claude-sonnet-4-6
      GEMINI_API_KEY set    → falls back to gemini-2.0-flash
      Neither set           → raises EnvironmentError
    """

    MAX_TOKENS = 4096

    def __init__(self):
        self.provider = self._detect_provider()
        self._init_client()
        print(f"{self.__class__.__name__} initialised [provider: {self.provider}]")

    # ── provider detection ────────────────────────────────────────

    def _detect_provider(self) -> str:
        if os.getenv("ANTHROPIC_API_KEY"):
            return "anthropic"
        elif os.getenv("GEMINI_API_KEY"):
            return "gemini"
        raise EnvironmentError(
            "No API key found. Add ANTHROPIC_API_KEY or "
            "GEMINI_API_KEY to your .env file."
        )

    def _init_client(self):
        if self.provider == "anthropic":
            from anthropic import Anthropic
            self.client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
            self.model  = "claude-sonnet-4-6"

        elif self.provider == "gemini":
            from google import genai
            self.client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
            self.model  = "gemini-2.0-flash"

    # ── AI call ───────────────────────────────────────────────────

    def call_claude(self, prompt: str, system: str = None) -> str:
        """
        Send prompt to AI provider. Returns plain text response.
        Method name kept as call_claude for consistency — works for
        both Anthropic and Gemini providers.
        """
        if self.provider == "anthropic":
            return self._call_anthropic(prompt, system)
        return self._call_gemini(prompt, system)

    def _call_anthropic(self, prompt: str, system: Optional[str]) -> str:
        kwargs = {
            "model":      self.model,
            "max_tokens": self.MAX_TOKENS,
            "temperature": 0,          # deterministic — same input → same output
            "messages":   [{"role": "user", "content": prompt}],
        }
        if system:
            kwargs["system"] = system
        response = self.client.messages.create(**kwargs)
        return response.content[0].text

    def _call_gemini(self, prompt: str, system: Optional[str]) -> str:
        from google.genai import types as genai_types
        full_prompt = f"{system}\n\n{prompt}" if system else prompt
        config      = genai_types.GenerateContentConfig(temperature=0)
        response    = self.client.models.generate_content(
            model=self.model,
            contents=full_prompt,
            config=config,
        )
        return response.text

    # ── input readers ─────────────────────────────────────────────

    def read_input(self, source_path: str) -> str:
        """
        Read test cases from any supported format.
        Returns plain text content for the AI prompt.

        Supported:
          .xlsx / .xls  — Excel workbook (requires openpyxl)
          .csv          — comma or tab separated
          .txt / .md    — plain text
          .json         — JSON array or object
        """
        path   = Path(source_path)
        suffix = path.suffix.lower()

        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")

        if suffix in (".xlsx", ".xls"):
            return self._read_excel(path)
        elif suffix == ".csv":
            return self._read_csv(path)
        elif suffix == ".json":
            return self._read_json(path)
        else:
            # .txt, .md, .py, anything else — read as text
            return path.read_text(encoding="utf-8")

    def _read_excel(self, path: Path) -> str:
        """
        Read all sheets from Excel file.
        Each row becomes a numbered test case in plain text.
        Headers are used as field names.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "pip install openpyxl to read Excel files. "
                "Add openpyxl to requirements.txt."
            )

        wb    = openpyxl.load_workbook(path, read_only=True)
        lines = [f"Source: {path.name}\n"]

        for sheet_name in wb.sheetnames:
            ws      = wb[sheet_name]
            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            lines.append(f"\n=== Sheet: {sheet_name} ===")
            headers = [str(h).strip() if h else f"Column{i}"
                       for i, h in enumerate(rows[0])]

            for i, row in enumerate(rows[1:], start=1):
                if all(cell is None for cell in row):
                    continue   # skip blank rows
                lines.append(f"\nTest Case {i}:")
                for header, value in zip(headers, row):
                    if value is not None:
                        lines.append(f"  {header}: {value}")

        wb.close()
        return "\n".join(lines)

    def _read_csv(self, path: Path) -> str:
        import csv
        lines = [f"Source: {path.name}\n"]
        with open(path, encoding="utf-8") as f:
            reader  = csv.DictReader(f)
            for i, row in enumerate(reader, start=1):
                lines.append(f"\nTest Case {i}:")
                for key, value in row.items():
                    if value:
                        lines.append(f"  {key}: {value}")
        return "\n".join(lines)

    def _read_json(self, path: Path) -> str:
        data = json.loads(path.read_text(encoding="utf-8"))
        return json.dumps(data, indent=2)

    # ── framework context reader ──────────────────────────────────

    def read_framework_context(self, project_root: str = ".") -> str:
        """
        Read the framework's own source files so the AI agent has full
        awareness of existing utilities, naming conventions, and patterns.

        Reads (in order, silently skips missing files):
          1. CLAUDE.md                  — naming rules, architecture, patterns
          2. utils/actions.py           — all available Actions methods
          3. utils/functions.py         — all available Functions methods
          4. utils/data_factory.py      — DataFactory methods
          5. context/test_context.py    — TestContext fields
          6. facade/*_facade.py         — existing Facade API (steps must call these)

        Returns a single context string to inject into agent system prompts.
        """
        root = Path(project_root)

        files_to_read = [
            ("FRAMEWORK CONVENTIONS (CLAUDE.md)",     root / "CLAUDE.md"),
            ("ACTIONS UTILITY (utils/actions.py)",     root / "utils" / "actions.py"),
            ("FUNCTIONS UTILITY (utils/functions.py)", root / "utils" / "functions.py"),
            ("DATA FACTORY (utils/data_factory.py)",   root / "utils" / "data_factory.py"),
            ("TEST CONTEXT (context/test_context.py)", root / "context" / "test_context.py"),
        ]

        # Dynamically include all existing facade files
        facade_dir = root / "facade"
        if facade_dir.exists():
            for f in sorted(facade_dir.glob("*_facade.py")):
                files_to_read.append((f"FACADE ({f.name})", f))

        sections = []
        for label, filepath in files_to_read:
            if not filepath.exists():
                continue
            content = filepath.read_text(encoding="utf-8")
            sections.append(
                f"\n{'='*60}\n"
                f"# {label}\n"
                f"# Path: {filepath}\n"
                f"{'='*60}\n"
                f"{content}"
            )

        if not sections:
            return ""

        header = (
            "\n\nFRAMEWORK CONTEXT\n"
            "=================\n"
            "The following are the EXISTING framework source files.\n"
            "You MUST use the utilities, naming conventions, and patterns shown here.\n"
            "Do NOT invent new helper names — use only what already exists.\n"
        )
        return header + "\n".join(sections)

    # ── output writer ─────────────────────────────────────────────

    def save_output(self, content: str, filepath: str) -> None:
        """Write agent output to file. Creates parent dirs automatically."""
        out = Path(filepath)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(content, encoding="utf-8")
        print(f"Saved → {filepath}")
